import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models.task import Task


# Column order must match what ExcelParser.REQUIRED_COLUMNS + OPTIONAL_COLUMNS expects
EXPORT_COLUMNS = [
    ('Site',              'site'),
    ('Category',          'category'),
    ('Product',           'product'),
    ('WP',                'wp'),
    ('WP-ID',             'wp_id'),
    ('Unit',              'unit'),
    ('Effort',            'effort'),
    ('Comment',           'comment'),
    ('Tuning Factor',     'tuning_factor'),
    ('Qty',               'qty'),
    ('Total',             'total'),
    ('Role',              'role'),
    ('Resource Category', 'resource_category'),
    ('Support Type',      'support_type'),
    ('SPC',               'spc'),
    ('Resource Name',     'resource_name'),
    ('Start Date',        'start_date'),
    ('End Date',          'end_date'),
    ('Completion %',      'completion_pct'),
]

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
CELL_FONT   = Font(name='Calibri', size=10)


def export_tasks_to_excel(tasks: List[Task]) -> io.BytesIO:
    """
    Export a list of Task ORM objects to an Excel workbook that matches
    the EffortEstimation import format, returned as an in-memory BytesIO stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'EffortEstimation'

    # Write header row
    headers = [col[0] for col in EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write task rows
    for task in tasks:
        row = []
        for _, field in EXPORT_COLUMNS:
            value = getattr(task, field, None)
            # Convert Decimal to float so openpyxl can serialize it
            if value is not None and hasattr(value, '__float__'):
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            row.append(value)
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = CELL_FONT

    # Auto-size columns
    for col_cells in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 50)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
